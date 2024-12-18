from datetime import datetime, timedelta

import pytz
from django.conf import settings
from django.core.mail import send_mail
from django.db.models import Count
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from .models import Robot


def export_data_to_excel_file():
    """
    Функция для создания Excel-файла, анализа и добавления в нее данных из БД, скачивания в браузере и загрузки в папку excel_files
    """
    tz = pytz.timezone("UTC")
    current_date = tz.localize(datetime.now())
    start_date = current_date - timedelta(days=7)
    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = (
        f'attachment; filename="report_{current_date.date()}.xlsx"'
    )

    wb = Workbook()

    queryset = (
        Robot.objects.values("model", "version")
        .annotate(total_count=Count("model"))
        .filter(created__range=(start_date, current_date))
    )

    for i in queryset:

        wb.create_sheet(title=f'{i["model"]}', index=0)
        ws = wb[i["model"]]
        if not ws.tables.get(i["model"], None):
            tab = Table(displayName=f"{i['model']}", ref="A1:C3")
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=True,
            )
            tab.tableStyleInfo = style
            ws.append(["Модель", "Версия", "Количество за неделю"])
            ws.add_table(tab)
        ws.append([i["model"], i["version"], i["total_count"]])

    sheets = wb.sheetnames
    queryset_model_list = [i["model"] for i in queryset]
    diff_list = list(set(sheets).difference(queryset_model_list))

    for d in diff_list:
        wb.remove(wb[d])
    wb.save(response)
    wb.save(f"excel_files/report_{current_date.date()}.xlsx")
    return response


def send_email(to_email, model, version):
    """
    Функция для отправки email-письма клиенту, сделавшему заказ на робота, при его появлении
    """
    send_mail(
        subject="Сообщение о наличии робота",
        message=f"""Добрый день!
                Недавно вы интересовались нашим роботом модели {model}, версии {version}.
                Этот робот теперь в наличии. Если вам подходит этот вариант - пожалуйста, свяжитесь с нами""",
        from_email=settings.EMAIL_HOST_USER,
        recipient_list=[to_email],
        fail_silently=False,
    )
